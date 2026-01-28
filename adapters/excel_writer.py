"""
Adapter: Write data to Excel files.
"""
from io import BytesIO
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook, load_workbook


class ExcelWriter:
    """Write tabular data to Excel."""
    
    def create_workbook(
        self,
        headers: List[str],
        rows: List[List[str]],
        sheet_name: str = "Sheet1"
    ) -> bytes:
        """
        Create a new Excel workbook.
        
        Args:
            headers: Column headers
            rows: Data rows
            sheet_name: Worksheet name
        
        Returns:
            Excel file bytes
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
    
    def fill_template(
        self,
        template_path: str,
        headers: List[str],
        rows: List[List[str]],
        sheet_name: Optional[str] = None
    ) -> bytes:
        """
        Fill an existing Excel template.
        
        Args:
            template_path: Path to template .xlsx
            headers: Data headers
            rows: Data rows
            sheet_name: Target sheet (uses active if None)
        
        Returns:
            Excel file bytes
        """
        path = Path(template_path)
        if not path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")
        
        workbook = load_workbook(template_path)
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        
        # Get template headers from first row
        template_headers = [cell.value for cell in sheet[1]]
        
        # Append rows aligned to template
        for row in rows:
            aligned_row = self._align_row(row, headers, template_headers)
            sheet.append(aligned_row)
        
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
    
    def append_to_file(
        self,
        file_path: str,
        headers: List[str],
        rows: List[List[str]],
        sheet_name: Optional[str] = None
    ) -> None:
        """
        Append rows to an existing Excel file.
        
        Args:
            file_path: Path to existing .xlsx
            headers: New data headers
            rows: New data rows
            sheet_name: Target sheet (uses active if None)
        """
        path = Path(file_path)
        
        if not path.exists():
            # Create new file
            content = self.create_workbook(headers, rows, sheet_name or "Sheet1")
            path.write_bytes(content)
            return
        
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        
        # Get existing headers
        existing_headers = [cell.value for cell in sheet[1]]
        
        # Append aligned rows
        for row in rows:
            aligned_row = self._align_row(row, headers, existing_headers)
            sheet.append(aligned_row)
        
        workbook.save(file_path)
    
    def _align_row(
        self,
        row: List[str],
        source_headers: List[str],
        target_headers: List[str]
    ) -> List[str]:
        """Align a row from source schema to target schema."""
        result = [""] * len(target_headers)
        for i, header in enumerate(source_headers):
            if header in target_headers and i < len(row):
                target_idx = target_headers.index(header)
                result[target_idx] = row[i]
        return result
