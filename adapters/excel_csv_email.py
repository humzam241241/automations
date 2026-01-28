"""
Adapter: Load emails from Excel or CSV files with auto-column detection.
"""
from pathlib import Path
from typing import Dict, List
import csv
from openpyxl import load_workbook


class ExcelCSVEmailAdapter:
    """Load emails from Excel or CSV files."""
    
    def __init__(self):
        pass
    
    def load_from_excel(self, file_path: str, sheet_name: str = None) -> tuple[List[str], List[Dict]]:
        """
        Load emails from Excel file.
        
        Args:
            file_path: Path to .xlsx file
            sheet_name: Sheet name (uses active if None)
        
        Returns:
            (headers, emails) - headers list and list of email dicts
        """
        path = Path(file_path)
        if not path.exists():
            return [], []
        
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        # Read headers from first row
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value))
        
        if not headers:
            return [], []
        
        # Normalize headers to standard email fields
        header_map = self._map_headers(headers)
        
        # Read data rows
        emails = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):  # Skip empty rows
                continue
            
            email_data = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    # Map to standard field if possible
                    standard_field = header_map.get(headers[i], headers[i])
                    email_data[standard_field] = str(value) if value is not None else ""
            
            # Ensure standard fields exist
            email = self._normalize_email(email_data)
            emails.append(email)
        
        wb.close()
        return headers, emails
    
    def load_from_csv(self, file_path: str) -> tuple[List[str], List[Dict]]:
        """
        Load emails from CSV file.
        
        Args:
            file_path: Path to .csv file
        
        Returns:
            (headers, emails) - headers list and list of email dicts
        """
        path = Path(file_path)
        if not path.exists():
            return [], []
        
        with path.open('r', encoding='utf-8', errors='ignore') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames if reader.fieldnames else []
            
            if not headers:
                return [], []
            
            # Normalize headers
            header_map = self._map_headers(headers)
            
            emails = []
            for row in reader:
                email_data = {}
                for header in headers:
                    standard_field = header_map.get(header, header)
                    email_data[standard_field] = row.get(header, "")
                
                email = self._normalize_email(email_data)
                emails.append(email)
        
        return headers, emails
    
    def _map_headers(self, headers: List[str]) -> Dict[str, str]:
        """
        Map various header names to standard email fields.
        
        Args:
            headers: List of header names from file
        
        Returns:
            Dict mapping file headers to standard fields
        """
        mapping = {}
        
        for header in headers:
            header_lower = header.lower().strip()
            
            # Subject mappings
            if header_lower in ['subject', 'title', 'topic', 'subj']:
                mapping[header] = 'subject'
            
            # From mappings
            elif header_lower in ['from', 'sender', 'from address', 'from email', 'sent by']:
                mapping[header] = 'from'
            
            # To mappings
            elif header_lower in ['to', 'recipient', 'to address', 'to email', 'sent to']:
                mapping[header] = 'to'
            
            # Date mappings
            elif header_lower in ['date', 'datetime', 'timestamp', 'received', 'sent date', 'received date']:
                mapping[header] = 'date'
            
            # Body mappings
            elif header_lower in ['body', 'message', 'content', 'text', 'description']:
                mapping[header] = 'body'
            
            # Attachments mappings
            elif header_lower in ['attachments', 'attachment', 'files', 'attached files']:
                mapping[header] = 'attachments'
            
            # Keep original if no mapping
            else:
                mapping[header] = header
        
        return mapping
    
    def _normalize_email(self, email_data: Dict) -> Dict:
        """
        Normalize email data to standard format.
        
        Args:
            email_data: Raw email data dict
        
        Returns:
            Normalized email dict with standard fields
        """
        return {
            "id": email_data.get("id", ""),
            "subject": email_data.get("subject", ""),
            "from": email_data.get("from", ""),
            "to": email_data.get("to", ""),
            "date": email_data.get("date", ""),
            "body": email_data.get("body", ""),
            "attachments": self._parse_attachments(email_data.get("attachments", "")),
            "attachment_text": [],
            # Include all other fields as-is
            **{k: v for k, v in email_data.items() if k not in ["id", "subject", "from", "to", "date", "body", "attachments"]}
        }
    
    def _parse_attachments(self, attachments_str: str) -> List[str]:
        """Parse attachment string into list."""
        if not attachments_str:
            return []
        
        # Try semicolon separator
        if ';' in attachments_str:
            return [a.strip() for a in attachments_str.split(';') if a.strip()]
        
        # Try comma separator
        if ',' in attachments_str:
            return [a.strip() for a in attachments_str.split(',') if a.strip()]
        
        # Single attachment
        return [attachments_str.strip()]
