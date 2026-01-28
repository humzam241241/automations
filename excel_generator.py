from io import BytesIO
from typing import List

from openpyxl import Workbook, load_workbook


def generate_excel(headers: List[str], row: List[str]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def fill_template(
    template_path: str,
    headers: List[str],
    row: List[str],
    summary: str,
) -> bytes:
    workbook = load_workbook(template_path)
    sheet = workbook.active

    template_headers = [cell.value for cell in sheet[1]]
    values = []
    for header in template_headers:
        if header in headers:
            values.append(row[headers.index(header)])
        else:
            values.append("")
    sheet.append(values)

    summary_row = [""] * len(template_headers)
    if template_headers:
        summary_row[0] = "SUMMARY"
    if "Summary" in template_headers:
        summary_row[template_headers.index("Summary")] = summary
    sheet.append(summary_row)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
