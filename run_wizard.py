"""
Offline wizard: Interactive profile creation and execution.
Detects Graph availability and offers local fallback.
"""
import json
import logging
import os
import sys
from pathlib import Path
from typing import Optional, Dict

import msal
import requests

from core.engine import ExecutionEngine
from core.profile_loader import ProfileLoader


def setup_logging():
    """Configure logging."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def get_token_cache_path() -> Path:
    """Get user-specific token cache path."""
    if os.name == "nt":  # Windows
        base = Path(os.getenv("LOCALAPPDATA", os.path.expanduser("~")))
    else:  # Unix-like
        base = Path(os.path.expanduser("~/.local/share"))
    
    cache_dir = base / "EmailAutomations"
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir / "token_cache.bin"


def test_graph_capabilities(token: str) -> Dict[str, bool]:
    """Test specific Graph API capabilities."""
    capabilities = {
        "me": False,
        "mail": False,
        "drive": False,
    }
    
    headers = {"Authorization": f"Bearer {token}"}
    
    # Test /me
    try:
        response = requests.get(
            "https://graph.microsoft.com/v1.0/me",
            headers=headers,
            timeout=10,
        )
        capabilities["me"] = response.status_code == 200
    except Exception:
        pass
    
    # Test mail access
    try:
        response = requests.get(
            "https://graph.microsoft.com/v1.0/me/mailFolders",
            headers=headers,
            timeout=10,
        )
        capabilities["mail"] = response.status_code == 200
    except Exception:
        pass
    
    # Test drive access
    try:
        response = requests.get(
            "https://graph.microsoft.com/v1.0/me/drive/root",
            headers=headers,
            timeout=10,
        )
        capabilities["drive"] = response.status_code == 200
    except Exception:
        pass
    
    return capabilities


def acquire_graph_token() -> Optional[str]:
    """Attempt to acquire Graph token via device code flow."""
    client_id = os.getenv("CLIENT_ID")
    authority_url = os.getenv("AUTHORITY")
    
    if not client_id or not authority_url:
        print("\n" + "="*60)
        print("MICROSOFT GRAPH AUTHENTICATION UNAVAILABLE")
        print("="*60)
        print("Missing environment variables: CLIENT_ID and/or AUTHORITY")
        print("Graph mode is disabled. You can only use local file inputs/outputs.")
        print("="*60)
        return None
    
    cache_path = get_token_cache_path()
    cache = msal.SerializableTokenCache()
    if cache_path.exists():
        cache.deserialize(cache_path.read_text(encoding="utf-8"))
    
    app = msal.PublicClientApplication(
        client_id=client_id,
        authority=authority_url,
        token_cache=cache,
    )
    
    scopes = ["Mail.Read", "Files.ReadWrite"]
    
    # Try silent first
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(scopes=scopes, account=accounts[0])
        if result and "access_token" in result:
            if cache.has_state_changed:
                cache_path.write_text(cache.serialize(), encoding="utf-8")
            return result["access_token"]
    
    # Device code flow
    print("\n" + "="*60)
    print("MICROSOFT GRAPH AUTHENTICATION")
    print("="*60)
    
    try:
        flow = app.initiate_device_flow(scopes=scopes)
        if "user_code" not in flow:
            print("Failed to start device flow.")
            return None
        
        print(flow["message"])
        result = app.acquire_token_by_device_flow(flow)
        
        if result and "access_token" in result:
            if cache.has_state_changed:
                cache_path.write_text(cache.serialize(), encoding="utf-8")
            return result["access_token"]
        else:
            print(f"Authentication failed: {result.get('error_description') if result else 'Unknown'}")
            return None
    except Exception as e:
        print(f"Authentication error: {e}")
        return None


def load_headers_from_template(template_path: str) -> list:
    """Load column headers from template .xlsx."""
    from openpyxl import load_workbook
    
    try:
        wb = load_workbook(template_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1] if cell.value]
        return headers
    except Exception as e:
        print(f"Failed to load template: {e}")
        return []


def wizard_create_profile(graph_available: bool, capabilities: Dict[str, bool]) -> dict:
    """Interactive wizard to create a profile."""
    print("\n" + "="*60)
    print("CREATE NEW PROFILE")
    print("="*60)
    
    profile = {}
    
    # Profile name
    profile["name"] = input("Profile name: ").strip()
    profile["description"] = input("Description (optional): ").strip()
    
    # Input source
    print("\nSelect input source:")
    if graph_available and capabilities.get("mail"):
        print("  1. Microsoft Graph (fetch from Outlook)")
    else:
        print("  1. Microsoft Graph (NOT AVAILABLE - missing mail permission)")
    print("  2. Local .eml files (no Graph required)")
    print("  3. Local CSV file (no Graph required)")
    
    source_choice = input("Choice [1-3]: ").strip()
    
    if source_choice == "1":
        if not graph_available or not capabilities.get("mail"):
            print("\n⚠ Graph mail access not available. Please choose option 2 or 3.")
            return wizard_create_profile(graph_available, capabilities)
        
        profile["input_source"] = "graph"
        
        # Email selection mode
        print("\nEmail selection mode:")
        print("  A. Folder + newest N messages")
        print("  B. Search query (from/subject/date/attachments)")
        sel_mode = input("Choice [A/B]: ").strip().upper()
        
        if sel_mode == "B":
            # Build search query
            print("\nBuild search query (leave blank to skip):")
            from_contains = input("  From contains: ").strip()
            subject_contains = input("  Subject contains: ").strip()
            after_date = input("  After date (YYYY-MM-DD): ").strip()
            has_attachments = input("  Has attachments? (y/n): ").strip().lower()
            
            search_parts = []
            if from_contains:
                search_parts.append(f"from:{from_contains}")
            if subject_contains:
                search_parts.append(f"subject:{subject_contains}")
            if has_attachments == "y":
                search_parts.append("hasAttachments:true")
            
            search_query = " AND ".join(search_parts) if search_parts else None
            
            profile["email_selection"] = {
                "search_query": search_query,
                "after_date": after_date if after_date else None,
                "newest_n": int(input("  Max results [25]: ").strip() or "25"),
            }
        else:
            # Folder mode
            folder_name = input("Outlook folder name (e.g., Inbox): ").strip()
            top = input("Number of newest messages [25]: ").strip() or "25"
            profile["email_selection"] = {
                "folder_name": folder_name,
                "newest_n": int(top),
                "search_query": None,
            }
    
    elif source_choice == "2":
        profile["input_source"] = "local_eml"
        directory = input("Directory path [./input_emails]: ").strip() or "./input_emails"
        pattern = input("File pattern [*.eml]: ").strip() or "*.eml"
        profile["email_selection"] = {
            "directory": directory,
            "pattern": pattern,
        }
    
    elif source_choice == "3":
        profile["input_source"] = "local_csv"
        csv_path = input("CSV file path [./emails.csv]: ").strip() or "./emails.csv"
        profile["email_selection"] = {
            "csv_path": csv_path,
        }
    
    else:
        print("Invalid choice. Defaulting to local .eml.")
        profile["input_source"] = "local_eml"
        profile["email_selection"] = {"directory": "./input_emails", "pattern": "*.eml"}
    
    # Schema
    print("\nDefine schema:")
    print("  A. Type column headers manually")
    print("  B. Load headers from template .xlsx")
    schema_choice = input("Choice [A/B]: ").strip().upper()
    
    if schema_choice == "B":
        template_path = input("Template .xlsx path: ").strip()
        headers = load_headers_from_template(template_path)
        if headers:
            print(f"Loaded {len(headers)} columns: {', '.join(headers)}")
            profile["schema"] = {
                "columns": [{"name": col, "type": "text"} for col in headers]
            }
        else:
            print("Failed to load template. Falling back to manual entry.")
            schema_choice = "A"
    
    if schema_choice == "A":
        print("Enter column headers (comma-separated, e.g., Subject,From,Date,Billing,IT):")
        columns_input = input("Columns: ").strip()
        columns = [c.strip() for c in columns_input.split(",") if c.strip()]
        
        profile["schema"] = {
            "columns": [{"name": col, "type": "text"} for col in columns]
        }
    
    # Rules
    print("\nDefine keyword/regex rules (leave blank to skip):")
    rules = []
    while True:
        col_name = input("Column name (or blank to finish): ").strip()
        if not col_name:
            break
        
        print(f"  Rule for column '{col_name}':")
        keywords_input = input("    Keywords (comma-separated): ").strip()
        keywords = [kw.strip() for kw in keywords_input.split(",") if kw.strip()]
        
        value = input("    Value if matched [Yes]: ").strip() or "Yes"
        
        # Advanced options
        word_boundary = input("    Use word boundaries? (y/n) [y]: ").strip().lower() != "n"
        
        print("    Search in: all, subject, body, attachments_text, attachments_names, from, to")
        search_in_input = input("    Search in [all]: ").strip() or "all"
        search_in = [s.strip() for s in search_in_input.split(",")]
        
        priority = input("    Priority (higher wins) [0]: ").strip()
        priority = int(priority) if priority.isdigit() else 0
        
        rule = {
            "column": col_name,
            "keywords": keywords,
            "value": value,
            "unmatched_value": "",
            "word_boundary": word_boundary,
            "search_in": search_in,
        }
        
        if priority > 0:
            rule["priority"] = priority
        
        rules.append(rule)
    
    profile["rules"] = rules
    
    # Output
    print("\nSelect output format:")
    print("  1. Excel")
    print("  2. CSV")
    format_choice = input("Choice [1-2]: ").strip()
    output_format = "excel" if format_choice == "1" else "csv"
    
    print("\nSelect output destination:")
    if graph_available and capabilities.get("drive"):
        print("  1. OneDrive")
    else:
        print("  1. OneDrive (NOT AVAILABLE - missing drive permission)")
    print("  2. Local file")
    dest_choice = input("Choice [1-2]: ").strip()
    
    if dest_choice == "1":
        if not graph_available or not capabilities.get("drive"):
            print("\n⚠ OneDrive access not available. Using local output.")
            dest_choice = "2"
    
    if dest_choice == "1":
        destination = "onedrive"
        onedrive_path = input("OneDrive folder path [/EmailReports]: ").strip() or "/EmailReports"
        profile["output"] = {
            "format": output_format,
            "destination": destination,
            "onedrive_path": onedrive_path,
            "filename_template": f"{profile['name']}_{{timestamp}}.{output_format if output_format == 'csv' else 'xlsx'}",
        }
    else:
        destination = "local"
        local_path = input("Local output directory [./output]: ").strip() or "./output"
        profile["output"] = {
            "format": output_format,
            "destination": destination,
            "local_path": local_path,
            "filename_template": f"{profile['name']}_{{timestamp}}.{output_format if output_format == 'csv' else 'xlsx'}",
        }
    
    return profile


def main():
    """Main wizard entry point."""
    setup_logging()
    
    print("="*60)
    print("EMAIL TO SPREADSHEET WIZARD")
    print("="*60)
    
    # Try to get Graph token
    print("\nChecking Microsoft Graph access...")
    token = acquire_graph_token()
    
    graph_available = False
    capabilities = {"me": False, "mail": False, "drive": False}
    
    if token:
        capabilities = test_graph_capabilities(token)
        graph_available = capabilities["me"]
        
        if graph_available:
            print("✓ Microsoft Graph connected")
            if capabilities["mail"]:
                print("  ✓ Mail access: OK")
            else:
                print("  ✗ Mail access: DENIED (cannot fetch emails)")
            
            if capabilities["drive"]:
                print("  ✓ OneDrive access: OK")
            else:
                print("  ✗ OneDrive access: DENIED (cannot upload to OneDrive)")
        else:
            print("✗ Microsoft Graph access denied")
            token = None
    else:
        print("✗ Microsoft Graph authentication unavailable")
    
    if not graph_available:
        print("\n" + "="*60)
        print("NO-GRAPH MODE")
        print("You can still use local .eml files or CSV inputs.")
        print("Output will be saved to local files only.")
        print("="*60)
    
    # Load existing profiles
    loader = ProfileLoader()
    existing_profiles = loader.list_profiles()
    
    if existing_profiles:
        print(f"\nExisting profiles: {', '.join(existing_profiles)}")
    
    print("\nOptions:")
    print("  1. Run existing profile")
    print("  2. Create new profile")
    print("  3. Exit")
    
    choice = input("\nChoice [1-3]: ").strip()
    
    if choice == "1":
        profile_name = input("Profile name: ").strip()
        profile = loader.load_profile(profile_name)
        
        if not profile:
            print(f"Profile '{profile_name}' not found.")
            sys.exit(1)
    
    elif choice == "2":
        profile = wizard_create_profile(graph_available, capabilities)
        
        # Validate
        errors = loader.validate_profile(profile)
        if errors:
            print("\nProfile validation errors:")
            for error in errors:
                print(f"  - {error}")
            sys.exit(1)
        
        # Save
        loader.save_profile(profile)
        print(f"\n✓ Profile '{profile['name']}' saved to profiles/{profile['name']}.json")
    
    else:
        print("Exiting.")
        sys.exit(0)
    
    # Check if profile requires Graph capabilities
    input_needs_mail = profile.get("input_source") == "graph"
    output_needs_drive = profile.get("output", {}).get("destination") == "onedrive"
    
    if input_needs_mail and not capabilities.get("mail"):
        print("\n✗ ERROR: This profile requires Mail.Read permission, but access was denied.")
        print("  Options:")
        print("    1. Use a different profile with local input")
        print("    2. Contact your IT admin to grant Mail.Read permission")
        sys.exit(1)
    
    if output_needs_drive and not capabilities.get("drive"):
        print("\n✗ ERROR: This profile requires Files.ReadWrite permission for OneDrive, but access was denied.")
        print("  Options:")
        print("    1. Use a different profile with local output")
        print("    2. Contact your IT admin to grant Files.ReadWrite permission")
        sys.exit(1)
    
    # Execute profile
    print(f"\nExecuting profile: {profile.get('name')}")
    engine = ExecutionEngine(access_token=token)
    
    try:
        result = engine.run_profile(profile, explain=True)
        print("\n" + "="*60)
        print("EXECUTION COMPLETE")
        print("="*60)
        print(json.dumps(result, indent=2))
    except Exception as e:
        logging.exception("Execution failed")
        print(f"\nERROR: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
