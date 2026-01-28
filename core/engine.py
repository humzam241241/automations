"""
Core execution engine - runs profiles with selected adapters.
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from adapters.csv_writer import CSVWriter
from adapters.excel_writer import ExcelWriter
from adapters.graph_email import GraphEmailAdapter
from adapters.local_email import LocalEmailAdapter
from adapters.onedrive_storage import OneDriveAdapter
from jobs.email_to_table import email_to_record
from jobs.ai_helper import AIHelper


class ExecutionEngine:
    """Execute profiles using configured adapters."""
    
    def __init__(self, access_token: Optional[str] = None, ai_helper: Optional[AIHelper] = None):
        """
        Initialize engine.
        
        Args:
            access_token: Microsoft Graph access token (required for Graph/OneDrive)
        """
        self.access_token = access_token
        self.ai_helper = ai_helper or AIHelper()
    
    def run_profile(self, profile: Dict, explain: bool = False) -> Dict:
        """
        Execute a profile configuration with pipeline support.
        
        Args:
            profile: Profile dict
            explain: If True, include rule match explanations
        
        Returns:
            Result dict with status, output_path, and optional explanation
        """
        logging.info(f"Running profile: {profile.get('name')}")
        
        # Get pipeline (default to email_to_table for backward compat)
        pipeline = profile.get("pipeline", ["email_to_table"])
        
        # 1. Load emails based on input_source
        emails = self._load_emails(profile)
        logging.info(f"Loaded {len(emails)} emails")
        
        if not emails:
            return {"status": "no_emails", "message": "No emails found"}
        
        # 2. Execute pipeline
        records = []
        explanations = []
        
        for job_name in pipeline:
            if job_name == "email_to_table":
                records, explanations = self._job_email_to_table(
                    emails, profile, explain
                )
            elif job_name == "append_to_master":
                records = self._job_append_to_master(records, profile)
            elif job_name == "excel_to_biready":
                records = self._job_excel_to_biready(records, profile)
            else:
                logging.warning(f"Unknown job: {job_name}")
        
        if not records:
            return {"status": "no_data", "message": "Pipeline produced no data"}
        
        # 3. Convert records to headers + rows for output
        headers, rows = self._records_to_table(records, profile)
        
        # 4. Check if should export to BI
        output_config = profile.get("output", {})
        should_export_bi = (
            output_config.get("destination") == "bi_dashboard" or
            output_config.get("also_export_bi", False) or
            profile.get("input_source") in ["excel_file", "csv_file"]
        )
        
        # 5. Write output
        output_result = self._write_output(profile, headers, rows)
        
        # 6. Generate BI dashboard if requested
        if should_export_bi:
            bi_result = self._export_bi_dashboard(profile, headers, rows)
            output_result["bi_dashboard"] = bi_result
        
        result = {
            "status": "success",
            "emails_processed": len(emails),
            "records_produced": len(records),
            "output": output_result,
            "headers": headers,
            "rows": rows,
            "profile_name": profile.get("name", "Unknown"),
        }
        
        if explain:
            result["explanations"] = explanations
        
        logging.info(f"Profile execution complete: {result}")
        return result
    
    def _job_email_to_table(
        self,
        emails: List[Dict],
        profile: Dict,
        explain: bool
    ) -> Tuple[List[Dict], List[Dict]]:
        """
        Execute email_to_table job.
        
        KEY FEATURES:
        1. If email has Excel/table rows, expand into multiple records
        2. If multiple order numbers found, create separate records per order
        3. Ensures each order number gets its own row with properly filled columns
        """
        schema = profile.get("schema", {})
        rules = profile.get("rules", [])
        use_synonyms = profile.get("use_synonyms", True)  # Default to True
        use_ai_flag = profile.get("use_ai", profile.get("ai_assist", False))
        ai_enabled = use_ai_flag and self.ai_helper and self.ai_helper.is_configured()
        ai_threshold = profile.get("ai_threshold", 4)
        
        records = []
        explanations = []
        
        for email_data in emails:
            # Check if email has Excel/table rows we should expand
            excel_rows = email_data.get("excel_rows", [])
            
            if excel_rows:
                # EXPAND: Create one output record per Excel/table row
                logging.info(f"Expanding email with {len(excel_rows)} table rows")
                for row_data in excel_rows:
                    # Merge email metadata with Excel row data
                    # Excel row data takes priority (it's the actual data we want!)
                    merged_data = dict(email_data)  # Copy email fields
                    merged_data.update(row_data)    # Add/override with Excel columns
                    
                    record, explanation = email_to_record(
                        merged_data, schema, rules, explain, use_synonyms, use_ai_flag, ai_threshold
                    )
                    if ai_enabled:
                        record, explanation = self._apply_ai_assist(
                            record, explanation, merged_data, ai_threshold
                        )
                    records.append(record)
                    if explain:
                        explanations.append(explanation)
            else:
                # Try to expand multiple order numbers into separate rows
                from jobs.email_to_table import expand_order_numbers_to_rows
                
                expanded = expand_order_numbers_to_rows(
                    email_data, schema, rules, explain, use_synonyms, use_ai_flag, ai_threshold
                )
                
                for record, explanation in expanded:
                    if ai_enabled:
                        record, explanation = self._apply_ai_assist(
                            record, explanation, email_data, ai_threshold
                        )
                    records.append(record)
                    if explain:
                        explanations.append(explanation)
        
        # Post-process: Fill missing columns from email context and merge tables
        records = self._fill_and_merge_records(records, emails, profile)
        
        return records, explanations

    def _apply_ai_assist(
        self,
        record: Dict[str, str],
        explanation: Dict[str, Dict],
        email_data: Dict,
        threshold: int
    ) -> Tuple[Dict[str, str], Dict[str, Dict]]:
        if not self.ai_helper or not self.ai_helper.is_configured():
            return record, explanation

        for column, meta in explanation.items():
            confidence = meta.get("confidence", 0)
            if confidence >= threshold:
                continue

            suggestion = self.ai_helper.suggest(column, email_data)
            if not suggestion or not suggestion.get("value"):
                continue

            new_value = suggestion["value"]
            record[column] = new_value
            explanation[column].update({
                "matched": True,
                "match_type": "ai_assist",
                "search_in": ["ai_assist"],
                "snippet": suggestion.get("snippet"),
                "confidence": suggestion.get("confidence", 6),
                "source": "ai_assist",
            })

        return record, explanation
    
    def _fill_and_merge_records(
        self,
        records: List[Dict],
        emails: List[Dict],
        profile: Dict
    ) -> List[Dict]:
        """
        Fill missing columns from email context and merge tables intelligently.
        
        This handles:
        1. Tables with different structures across emails
        2. Missing columns in some tables
        3. Merging by index column (Work Order number)
        4. Filling gaps from email body/text when columns aren't in tables
        """
        if not records:
            return records
        
        try:
            from jobs.table_merger import merge_tables, fill_missing_columns_from_context
        except ImportError:
            logging.warning("table_merger not available, skipping merge")
            return records
        
        schema = profile.get("schema", {})
        schema_columns = schema.get("columns", [])
        schema_column_names = [col.get("name", "") for col in schema_columns]
        use_synonyms = profile.get("use_synonyms", True)
        index_column = profile.get("index_column")  # e.g., "order", "wo number"
        
        # Step 1: Fill missing columns from email context
        # Map records back to their source emails (if possible)
        # For now, we'll fill from all emails (less precise but works)
        filled_records = []
        
        for record in records:
            # Try to find source email (if record has _source_file or we can match by order)
            source_email = None
            
            # Look for email that might have this data
            order_val = None
            for col_name in ["order", "Order", "work order", "Work Order"]:
                if col_name in record and record[col_name]:
                    order_val = str(record[col_name]).strip()
                    break
            
            if order_val:
                # Try to find email containing this order number
                for email in emails:
                    email_text = f"{email.get('subject', '')} {email.get('body', '')} {' '.join(email.get('attachment_text', []))}"
                    if order_val in email_text:
                        source_email = email
                        break
            
            # If no specific email found, use first email as fallback
            if not source_email and emails:
                source_email = emails[0]
            
            # Fill missing columns from email context
            if source_email:
                filled_record = fill_missing_columns_from_context(
                    record, source_email, schema_columns, use_synonyms
                )
            else:
                filled_record = record
            
            filled_records.append(filled_record)
        
        # Step 2: Merge tables intelligently
        merged = merge_tables(
            filled_records,
            index_column=index_column,
            schema_columns=schema_column_names
        )
        
        logging.info(f"Filled and merged {len(records)} records into {len(merged)} final records")
        return merged
    
    def _job_append_to_master(
        self,
        records: List[Dict],
        profile: Dict
    ) -> List[Dict]:
        """Execute append_to_master job."""
        try:
            from jobs.append_to_master import append_rows
            
            master_config = profile.get("master_dataset", {})
            master_path = master_config.get("path")
            dedupe_column = master_config.get("deduplicate_column")
            
            if not master_path:
                logging.warning("No master_dataset.path in profile, skipping append")
                return records
            
            # Load existing master if exists
            existing_records = self._load_master_records(master_path)
            
            # Convert to headers+rows for append_rows function
            if existing_records:
                existing_headers = list(existing_records[0].keys())
                existing_rows = [list(r.values()) for r in existing_records]
            else:
                existing_headers = []
                existing_rows = []
            
            new_headers = list(records[0].keys()) if records else []
            new_rows = [list(r.values()) for r in records]
            
            merged_headers, merged_rows = append_rows(
                existing_headers,
                existing_rows,
                new_headers,
                new_rows,
                dedupe_column
            )
            
            # Convert back to records
            merged_records = []
            for row in merged_rows:
                record = {h: v for h, v in zip(merged_headers, row)}
                merged_records.append(record)
            
            # Save master
            self._save_master_records(master_path, merged_records)
            
            return merged_records
        
        except ImportError:
            logging.warning("append_to_master job not available")
            return records
    
    def _job_excel_to_biready(
        self,
        records: List[Dict],
        profile: Dict
    ) -> List[Dict]:
        """Execute excel_to_biready job."""
        try:
            from jobs.excel_to_biready import apply_bi_transformations
            
            transformations = profile.get("bi_transformations", [])
            if not transformations:
                return records
            
            # Convert to headers+rows
            headers = list(records[0].keys()) if records else []
            rows = [list(r.values()) for r in records]
            
            # Apply transformations
            new_headers, new_rows = apply_bi_transformations(
                headers, rows, transformations
            )
            
            # Convert back to records
            new_records = []
            for row in new_rows:
                record = {h: v for h, v in zip(new_headers, row)}
                new_records.append(record)
            
            return new_records
        
        except ImportError:
            logging.warning("excel_to_biready job not available")
            return records
    
    def _records_to_table(
        self,
        records: List[Dict],
        profile: Dict
    ) -> tuple[List[str], List[List[str]]]:
        """Convert records to headers + rows, preserving schema order."""
        if not records:
            return [], []
        
        # Get headers from profile schema to preserve order
        schema = profile.get("schema", {})
        if schema and schema.get("columns"):
            headers = [
                col.get("name", col.get("header", ""))
                for col in schema.get("columns", [])
            ]
        else:
            # Fallback: use keys from first record
            headers = list(records[0].keys())
        
        # Build rows aligned to headers
        rows = []
        for record in records:
            row = [record.get(h, "") for h in headers]
            rows.append(row)
        
        return headers, rows
    
    def _load_master_records(self, path: str) -> List[Dict]:
        """Load existing master dataset."""
        from pathlib import Path
        import csv
        
        file_path = Path(path)
        if not file_path.exists():
            return []
        
        records = []
        if path.endswith(".csv"):
            with file_path.open("r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                records = list(reader)
        elif path.endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                record = {h: v for h, v in zip(headers, row)}
                records.append(record)
        
        return records
    
    def _save_master_records(self, path: str, records: List[Dict]):
        """Save master dataset."""
        if not records:
            return
        
        headers = list(records[0].keys())
        rows = [list(r.values()) for r in records]
        
        if path.endswith(".csv"):
            writer = CSVWriter()
            content = writer.create_csv(headers, rows)
            Path(path).write_bytes(content)
        elif path.endswith(".xlsx"):
            writer = ExcelWriter()
            content = writer.create_workbook(headers, rows)
            Path(path).write_bytes(content)
    
    def _load_emails(self, profile: Dict) -> List[Dict]:
        """Load emails based on profile input_source."""
        input_source = profile.get("input_source")
        email_selection = profile.get("email_selection", {})
        
        if input_source == "graph":
            if not self.access_token:
                raise RuntimeError("Graph access token required for input_source='graph'")
            
            adapter = GraphEmailAdapter(self.access_token)
            folder_name = email_selection.get("folder_name")
            search_query = email_selection.get("search_query")
            top = email_selection.get("newest_n", 25)
            
            folder_id = None
            if folder_name:
                folder_id = adapter.get_folder_id(folder_name)
            
            return adapter.fetch_messages(folder_id, search_query, top)
        
        elif input_source == "local_eml":
            adapter = LocalEmailAdapter(extract_attachments=True)
            directory = email_selection.get("directory", "./input_emails")
            pattern = email_selection.get("pattern", "*.eml")
            file_paths = email_selection.get("file_paths", [])
            
            # Filter out empty paths
            file_paths = [p for p in file_paths if p and p.strip()]
            
            if file_paths:
                # Use the new load_from_files which handles mixed types
                return adapter.load_from_files(file_paths)
            else:
                return adapter.scan_directory(directory, pattern)
        
        elif input_source == "local_csv":
            adapter = LocalEmailAdapter()
            csv_path = email_selection.get("csv_path", "./emails.csv")
            return adapter.load_from_csv(csv_path)
        
        elif input_source == "excel_file":
            from adapters.excel_csv_email import ExcelCSVEmailAdapter
            adapter = ExcelCSVEmailAdapter()
            
            # Support multiple files
            file_paths = email_selection.get("file_paths", [])
            file_path = email_selection.get("file_path", "")
            
            # Use file_paths if available, else single file_path
            if not file_paths and file_path:
                file_paths = [file_path]
            
            file_paths = [p for p in file_paths if p and p.strip()]
            
            if not file_paths:
                raise ValueError("Excel file path is required")
            
            # Load from multiple files
            all_emails = []
            all_headers = []
            
            for fp in file_paths:
                fp = fp.strip()
                ext = fp.lower().split('.')[-1]
                
                if ext in ['pdf', 'docx', 'txt']:
                    # Use LocalEmailAdapter for document files
                    doc_adapter = LocalEmailAdapter(extract_attachments=True)
                    all_emails.extend(doc_adapter.load_from_files([fp]))
                elif ext == 'csv':
                    headers, emails = adapter.load_from_csv(fp)
                    if headers:
                        all_headers = headers
                    all_emails.extend(emails)
                else:
                    headers, emails = adapter.load_from_excel(fp)
                    if headers:
                        all_headers = headers
                    all_emails.extend(emails)
            
            # If auto-detect enabled, update schema
            if profile.get("auto_detect_columns") and all_headers:
                profile["schema"] = {
                    "columns": [{"name": h, "type": "text"} for h in all_headers]
                }
                logging.info(f"Auto-detected {len(all_headers)} columns from files")
            
            return all_emails
        
        elif input_source == "csv_file":
            from adapters.excel_csv_email import ExcelCSVEmailAdapter
            adapter = ExcelCSVEmailAdapter()
            
            # Support multiple files
            file_paths = email_selection.get("file_paths", [])
            file_path = email_selection.get("file_path", "")
            
            if not file_paths and file_path:
                file_paths = [file_path]
            
            file_paths = [p for p in file_paths if p and p.strip()]
            
            if not file_paths:
                raise ValueError("CSV file path is required")
            
            # Load from multiple files
            all_emails = []
            all_headers = []
            
            for fp in file_paths:
                fp = fp.strip()
                ext = fp.lower().split('.')[-1]
                
                if ext in ['pdf', 'docx', 'txt']:
                    doc_adapter = LocalEmailAdapter(extract_attachments=True)
                    all_emails.extend(doc_adapter.load_from_files([fp]))
                elif ext in ['xlsx', 'xls']:
                    headers, emails = adapter.load_from_excel(fp)
                    if headers:
                        all_headers = headers
                    all_emails.extend(emails)
                else:
                    headers, emails = adapter.load_from_csv(fp)
                    if headers:
                        all_headers = headers
                    all_emails.extend(emails)
            
            # If auto-detect enabled, update schema
            if profile.get("auto_detect_columns") and all_headers:
                profile["schema"] = {
                    "columns": [{"name": h, "type": "text"} for h in all_headers]
                }
                logging.info(f"Auto-detected {len(all_headers)} columns from CSV files")
            
            return all_emails
        
        else:
            raise ValueError(f"Unknown input_source: {input_source}")
    
    def _write_output(self, profile: Dict, headers: List[str], rows: List[List[str]]) -> Dict:
        """Write output based on profile configuration."""
        output = profile.get("output", {})
        output_format = output.get("format", "excel")
        destination = output.get("destination", "local")
        
        # Generate filename
        filename_template = output.get("filename_template", "output_{timestamp}.xlsx")
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = filename_template.replace("{timestamp}", timestamp)
        
        # Generate file content
        if output_format == "excel":
            writer = ExcelWriter()
            template_path = output.get("template_path")
            if template_path:
                content = writer.fill_template(template_path, headers, rows)
            else:
                content = writer.create_workbook(headers, rows)
        elif output_format == "csv":
            writer = CSVWriter()
            content = writer.create_csv(headers, rows)
        else:
            raise ValueError(f"Unknown output format: {output_format}")
        
        # Write to destination
        if destination == "onedrive":
            if not self.access_token:
                raise RuntimeError("Graph access token required for destination='onedrive'")
            
            onedrive_path = output.get("onedrive_path", "/EmailReports")
            adapter = OneDriveAdapter(self.access_token)
            
            content_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                if output_format == "excel"
                else "text/csv"
            )
            
            result = adapter.upload_file(onedrive_path, filename, content, content_type)
            return {
                "destination": "onedrive",
                "path": f"{onedrive_path}/{filename}",
                "web_url": result.get("webUrl") if result else None,
            }
        
        elif destination == "local":
            local_path = output.get("local_path", "./output")
            Path(local_path).mkdir(parents=True, exist_ok=True)
            
            output_file = Path(local_path) / filename
            output_file.write_bytes(content)
            
            return {
                "destination": "local",
                "path": str(output_file.absolute()),
            }
        
        else:
            raise ValueError(f"Unknown destination: {destination}")
    
    def _export_bi_dashboard(self, profile: Dict, headers: List[str], rows: List[List[str]]) -> Dict:
        """Export to BI dashboard."""
        try:
            from bi_dashboard_export import generate_dashboard, save_dashboard
            import webbrowser
            
            profile_name = profile.get("name", "Dashboard")
            html = generate_dashboard(headers, rows, profile_name)
            
            # Save dashboard
            output_dir = Path("./output/dashboards")
            output_dir.mkdir(parents=True, exist_ok=True)
            
            path = save_dashboard(html, output_dir)
            
            # Open in browser
            webbrowser.open(f"file:///{path}")
            
            logging.info(f"BI Dashboard saved: {path}")
            
            return {
                "enabled": True,
                "path": path,
                "opened_in_browser": True
            }
        
        except Exception as e:
            logging.error(f"Failed to export BI dashboard: {e}")
            return {
                "enabled": False,
                "error": str(e)
            }